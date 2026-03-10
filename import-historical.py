#!/usr/bin/env python3
"""
Import historical paddle tennis match data from WBP season Excel files.

Single season:
  python3 import-historical.py --file "/path/to/History/WBP 24-25.xlsx"
  python3 import-historical.py --file "/path/to/History/WBP 24-25.xlsx" --season 2024-25

All seasons in a directory (processes both .xlsx and .xls, skips golf files):
  python3 import-historical.py --dir "/path/to/History"
  python3 import-historical.py --dir "/path/to/History" --skip 2024-25

Single-season output:   historical-members.sql + historical-data.sql
Multi-season output:    historical-members-all.sql + historical-data-all.sql

Run BOTH sql files in order in the Supabase SQL editor:
  1. historical-members[-all].sql  (adds new/guest players)
  2. historical-data[-all].sql     (deletes old season data, inserts fresh matches)
"""

import openpyxl
try:
    import xlrd
    HAS_XLRD = True
except ImportError:
    HAS_XLRD = False
import re
import sys
import os
import glob
import argparse
from datetime import datetime, timedelta

# ── Season start dates (Week 1 date, keyed by season label) ─────────────────
# These are estimated Monday start dates; close enough for historical purposes.

SEASON_STARTS = {
    '2024-25': datetime(2024, 10, 16),
    '2023-24': datetime(2023, 10, 17),
    '2022-23': datetime(2022, 10, 18),
    '2021-22': datetime(2021, 10, 19),
    '2020-21': datetime(2020, 10, 21),
    '2019-20': datetime(2019, 10, 15),
    '2018-19': datetime(2018, 10, 16),
    '2017-18': datetime(2017, 10, 17),
    '2016-17': datetime(2016, 10, 19),
    '2015-16': datetime(2015, 10, 20),
    '2014-15': datetime(2014, 10, 21),
    '2013-14': datetime(2013, 10, 15),
    '2012-13': datetime(2012, 10, 17),
    '2011-12': datetime(2011, 10, 18),
    # .xls seasons (these were already correct)
    '2010-11': datetime(2010, 10, 11),
    '2009-10': datetime(2009, 10, 12),
    '2008-09': datetime(2008, 10, 13),
    '2007-08': datetime(2007, 10, 15),
    '2006-07': datetime(2006, 10, 16),
}

# ── Score / team name detection ───────────────────────────────────────────────

# Headers that sometimes bleed into the member list column
_HEADER_RE = re.compile(r'^(week|court|average|name|position|record|total|subbed|matches)\s*\d*$', re.IGNORECASE)

def is_valid_member_name(v):
    """Return True if v looks like a real player name (has letters, not a header)."""
    if not isinstance(v, str):
        return False
    s = v.strip()
    return bool(re.search(r'[a-zA-Z]', s)) and not _HEADER_RE.match(s)

def is_team_name(v):
    """
    Return True if v looks like 'LastName/LastName' — has letters on BOTH sides of '/'.
    E.g. "Jonson/Richards" → True.  "6/2/6" → False.  "Rained Out" → False.
    """
    if not v or not isinstance(v, str):
        return False
    parts = str(v).split('/')
    return len(parts) >= 2 and all(re.search(r'[a-zA-Z]', p) for p in parts[:2])

def is_score_str(v):
    """
    Return True if v looks like a score string — only digits, '/', '(', ')' and must contain '/'.
    E.g. "6/2/6", "6(4)/7", "1/1" → True.  "Jonson/Richards", "Rained Out" → False.
    """
    if not v:
        return False
    s = str(v).strip()
    return '/' in s and bool(re.match(r"^[\d/()\'\s]+$", s))

# Names to treat as "no player" (byes, placeholders, etc.)
_INVALID_NAMES = {'NA', 'N/A', 'TBD', 'BYE', ''}

def normalize_player_name(name):
    """
    Clean up a raw player name from the spreadsheet:
    - Strip substitute markers: "Harris(sub)" → "Harris"
    - Normalize whitespace
    - Return None for known-invalid placeholders
    """
    if not name:
        return None
    name = str(name).strip()
    # Strip any parenthetical containing 'sub' (handles many annotation styles):
    #   "(sub)", "(subbed)", "(Sub for Laird)", "(Craigin Sub)", "(Duane Biasi sub)"
    name = re.sub(r'\s*\([^)]*\bsub\b[^)]*\)\s*', '', name, flags=re.IGNORECASE).strip()
    # Strip trailing notes after comma: "Smith, out Week 3" → "Smith"
    name = re.split(r'\s*,\s*(?:out|dnf|rained)', name, maxsplit=1, flags=re.IGNORECASE)[0].strip()
    name = re.sub(r'\s+', ' ', name)
    if name.upper() in _INVALID_NAMES:
        return None
    return name if name else None

def get_team_and_score(row, col):
    """
    Given a row values tuple and starting column index, return (team_str, score_str).

    Handles two layouts found in WBP files:
      Standard (2015-16+): row[col]=team_name, row[col+1]=score
      Inverted (2011-14):   row[col]=score,     row[col+1]=team_name

    Returns (None, None) if no valid team is found at this column.
    """
    val0 = row[col]     if col     < len(row) else None
    val1 = row[col + 1] if col + 1 < len(row) else None

    if is_team_name(val0):
        return str(val0), val1          # standard format
    elif is_score_str(val0) and is_team_name(val1):
        return str(val1), val0          # inverted format
    return None, None                   # rained out, blank, etc.

def parse_score(s):
    """
    Parse a score string like '6/2/6', '6/6', '6(4)/7/6', "6/2'6", '/4/6/1'.
    Returns list of ints, or [] if unparseable.
    """
    if not s:
        return []
    s = str(s).strip()
    if not s:
        return []
    s = s.lstrip('/')                        # leading slash typo
    s = re.sub(r'\(\d+\)', '', s)           # strip tiebreak: 6(4) -> 6
    s = s.replace("'", '/')                  # apostrophe typo: 6/2'6 -> 6/2/6
    parts = [p.strip() for p in s.split('/') if p.strip()]
    result = []
    try:
        for p in parts:
            n = int(p)
            # Detect merged set scores > 20 where both digits are valid set values (0-7).
            # e.g. "66" → likely "6/6" typo → expand to [6, 6].
            # Numbers ≤ 20 are kept as-is (could be a 10-pt tiebreak: 11/9, 10/8, etc.)
            if n > 20:
                tens, ones = n // 10, n % 10
                if tens <= 7 and ones <= 7:
                    result.extend([tens, ones])
                    continue
            result.append(n)
    except ValueError:
        return []
    return result

def null_or_int(scores, idx):
    """Return scores[idx] as string int, or 'NULL'."""
    if idx < len(scores):
        return str(scores[idx])
    return 'NULL'

# ── Excel parsing ─────────────────────────────────────────────────────────────

def parse_season_file(xlsx_path, season_label):
    """
    Parse a WBP season Excel file.
    Returns (matches, season_members, use_named_members) where:
      matches            = list of match dicts
      season_members     = set of last names for this season
      use_named_members  = True if member list came from D4:D20 (named)
                           False if derived from match data (pre-2017 rating files)
    """
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb['Scoring']

    # Read member list: col D (index 3), rows 4-20
    raw_vals = []
    for row in ws.iter_rows(min_row=4, max_row=20, min_col=4, max_col=4, values_only=True):
        if row[0] is not None:
            raw_vals.append(row[0])

    # Pre-2017 files store numerical ratings in D4:D20, not names.
    # Detect by counting how many values look like real player names.
    name_count = sum(1 for v in raw_vals if is_valid_member_name(v))
    use_named_members = name_count >= 8

    if use_named_members:
        season_members = {str(v).strip() for v in raw_vals if is_valid_member_name(v)}
        print(f"  Member list from D4:D20 ({len(season_members)}): {sorted(season_members)}")
    else:
        season_members = None   # will be derived from match data below
        print(f"  Member list: numerical ratings detected — will derive from match data")

    # Find court rows by scanning col B (index 1) for "Court N"
    court_rows = []  # list of (top_row_values, bot_row_values, court_name)
    all_rows = list(ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True))

    i = 0
    while i < len(all_rows):
        row = all_rows[i]
        col_b = row[1] if len(row) > 1 else None
        if col_b and str(col_b).strip().startswith('Court'):
            court_name = str(col_b).strip()
            top_row = row
            vs_row  = all_rows[i + 1] if i + 1 < len(all_rows) else []
            bot_row = all_rows[i + 2] if i + 2 < len(all_rows) else []
            court_rows.append((top_row, bot_row, court_name))
            i += 3
            continue
        i += 1

    print(f"  Courts found: {[c for _, _, c in court_rows]}")

    # Week columns: 3, 6, 9, ... up to max_column
    week_cols = list(range(3, ws.max_column, 3))

    matches = []
    for top_row, bot_row, court_name in court_rows:
        for col in week_cols:
            # Smart detection: handles both standard (team/score) and inverted (score/team) layouts
            top_team_str, top_score_str = get_team_and_score(top_row, col)
            if not top_team_str:
                continue   # rained out, blank, or non-player data

            bot_team_str, bot_score_str = get_team_and_score(bot_row, col)

            # Parse and normalize team names
            top_parts = [normalize_player_name(p) for p in top_team_str.split('/')]
            top_parts = [p for p in top_parts if p]   # drop None/empty
            bot_parts = []
            if bot_team_str:
                bot_parts = [normalize_player_name(p) for p in bot_team_str.split('/')]
                bot_parts = [p for p in bot_parts if p]

            if len(top_parts) != 2:
                continue  # skip malformed or incomplete entries
            if len(bot_parts) != 2:
                continue  # skip bye weeks / missing opponent

            top_scores = parse_score(top_score_str)
            bot_scores = parse_score(bot_score_str)

            week_num = (col - 3) // 3 + 1

            matches.append({
                'season':        season_label,
                'week_num':      week_num,
                'week_label':    f'Week {week_num}',
                'court':         court_name,
                'team1_player1': top_parts[0],
                'team1_player2': top_parts[1],
                'team2_player1': bot_parts[0] if len(bot_parts) > 0 else None,
                'team2_player2': bot_parts[1] if len(bot_parts) > 1 else None,
                'team1_scores':  top_scores,
                'team2_scores':  bot_scores,
            })

    # For pre-2017 files: derive season_members from all participants
    if season_members is None:
        season_members = set()
        for m in matches:
            for name in [m['team1_player1'], m['team1_player2'], m['team2_player1'], m['team2_player2']]:
                if name:
                    season_members.add(name)
        print(f"  Derived {len(season_members)} members from match data: {sorted(season_members)}")

    return matches, season_members, use_named_members

# ── .xls parsing (2006-07 through 2010-11) ────────────────────────────────────

def parse_xls_file(xls_path, season_label):
    """
    Parse a WBP season .xls file using xlrd.
    Structure (Scoring sheet):
      - Member names: col 2 (0-indexed), rows 3-21 (0-indexed)
      - Court rows: col 1 contains 'Court N'
      - Week columns: 3, 6, 9, ... (step 3)
      - Match format: score at col N, team name at col N+1 (inverted layout)
    Returns (matches, season_members, use_named_members).
    """
    if not HAS_XLRD:
        print('ERROR: xlrd not installed. Run: pip3 install xlrd')
        return [], set(), True

    wb = xlrd.open_workbook(xls_path)
    ws = wb.sheet_by_name('Scoring')

    # Extract member list from col 2, rows 3-21 (0-indexed)
    season_members = set()
    for r in range(3, 22):
        if r >= ws.nrows:
            break
        v = ws.cell_value(r, 2)
        v_str = str(v).strip() if v else ''
        if is_valid_member_name(v_str):
            season_members.add(v_str)

    print(f"  Member list from col C ({len(season_members)}): {sorted(season_members)}")

    # Find court rows by scanning col 1 for "Court" keyword
    court_top_rows = []
    for r in range(ws.nrows):
        v = str(ws.cell_value(r, 1)).strip()
        if re.match(r'Court\s*\d*', v, re.IGNORECASE) and v.lower() != 'court':
            court_top_rows.append(r)

    print(f"  Courts found: {len(court_top_rows)}")

    # Week columns: 3, 6, 9, ...
    week_cols = list(range(3, ws.ncols - 1, 3))

    matches = []
    for court_idx, top_r in enumerate(court_top_rows):
        bot_r = top_r + 2   # skip 'vs' row at top_r + 1
        if bot_r >= ws.nrows:
            continue

        # Build row lists compatible with get_team_and_score()
        top_row = [ws.cell_value(top_r, c) if c < ws.ncols else None for c in range(ws.ncols)]
        bot_row = [ws.cell_value(bot_r, c) if c < ws.ncols else None for c in range(ws.ncols)]

        for col in week_cols:
            top_team_str, top_score_str = get_team_and_score(top_row, col)
            if not top_team_str:
                continue

            bot_team_str, bot_score_str = get_team_and_score(bot_row, col)

            top_parts = [normalize_player_name(p) for p in top_team_str.split('/')]
            top_parts = [p for p in top_parts if p]
            bot_parts = []
            if bot_team_str:
                bot_parts = [normalize_player_name(p) for p in bot_team_str.split('/')]
                bot_parts = [p for p in bot_parts if p]

            if len(top_parts) != 2:
                continue
            if len(bot_parts) != 2:
                continue

            top_scores = parse_score(top_score_str)
            bot_scores = parse_score(bot_score_str)

            week_num = (col - 3) // 3 + 1

            matches.append({
                'season':        season_label,
                'week_num':      week_num,
                'week_label':    f'Week {week_num}',
                'court':         f'Court {court_idx + 1}',
                'team1_player1': top_parts[0],
                'team1_player2': top_parts[1],
                'team2_player1': bot_parts[0],
                'team2_player2': bot_parts[1],
                'team1_scores':  top_scores,
                'team2_scores':  bot_scores,
            })

    # Fall back to deriving members from match data if list was too sparse
    if len(season_members) < 4:
        season_members = set()
        for m in matches:
            for name in [m['team1_player1'], m['team1_player2'],
                         m['team2_player1'], m['team2_player2']]:
                if name:
                    season_members.add(name)
        print(f"  Derived {len(season_members)} members from match data")

    # Return use_named_members=False so the batch mode inserts ALL participants.
    # This ensures early players (Laird, Zejda, DeGulis, etc.) that predate
    # the 2011-12 import get added to the DB, with is_guest=False for members
    # and is_guest=True for guests.
    # We handle the guest flag ourselves below so caller gets a full list.
    return matches, season_members, False  # False → batch inserts all players


# ── SQL generation ────────────────────────────────────────────────────────────

def member_lookup(name):
    """Generate a subquery to look up a member ID by last name."""
    if not name:
        return 'NULL'
    safe = name.replace("'", "''")
    return f"(SELECT id FROM members WHERE LOWER(full_name) = LOWER('{safe}') LIMIT 1)"

def week_to_date(season_label, week_num):
    start = SEASON_STARTS.get(season_label)
    if not start:
        return "'2000-01-01'"
    d = start + timedelta(weeks=week_num - 1)
    return f"'{d.strftime('%Y-%m-%d')}'"

def name_to_email(name):
    """Generate a placeholder email for historical members."""
    safe = name.lower().replace(' ', '.').replace("'", '')
    return f"{safe}@historical.local"

def generate_members_sql(all_new_members):
    """
    Generate SQL to insert new members.
    all_new_members: list of (name, is_guest) tuples
    """
    lines = [
        '-- =============================================',
        '-- historical-members-all.sql',
        '-- Run FIRST in Supabase SQL editor.',
        '-- Adds historical players not yet in the members table.',
        '-- =============================================',
        '',
    ]
    if all_new_members:
        for name, is_guest in sorted(all_new_members, key=lambda x: x[0]):
            safe = name.replace("'", "''")
            email = name_to_email(name)
            guest_str = 'true' if is_guest else 'false'
            lines.append(
                f"INSERT INTO members (full_name, email, is_guest, is_admin)"
                f" VALUES ('{safe}', '{email}', {guest_str}, false)"
                f" ON CONFLICT DO NOTHING;"
            )
        lines.append('')
        lines.append(f'-- {len(all_new_members)} player(s) added')
    else:
        lines.append('-- No new players to add.')
    return '\n'.join(lines)

def generate_data_sql(all_season_matches):
    """
    Generate SQL to delete and re-insert matches for multiple seasons.
    all_season_matches: list of (season_label, matches) tuples
    """
    lines = [
        '-- =============================================',
        '-- historical-data-all.sql',
        '-- Run SECOND in Supabase SQL editor (after historical-members-all.sql).',
        f'-- Covers {len(all_season_matches)} season(s).',
        '-- =============================================',
        '',
    ]
    total = 0
    for season_label, matches in all_season_matches:
        lines.append(f"DELETE FROM matches WHERE season = '{season_label}';")
        lines.append(f'')
        lines.append(f'-- === Season {season_label} ({len(matches)} matches) ===')
        lines.append('')
        for m in matches:
            t1p1 = member_lookup(m['team1_player1'])
            t1p2 = member_lookup(m['team1_player2'])
            t2p1 = member_lookup(m['team2_player1'])
            t2p2 = member_lookup(m['team2_player2'])
            date_str = week_to_date(m['season'], m['week_num'])

            s1 = m['team1_scores']
            s2 = m['team2_scores']

            t1s1 = null_or_int(s1, 0)
            t1s2 = null_or_int(s1, 1)
            t1s3 = null_or_int(s1, 2)
            t2s1 = null_or_int(s2, 0)
            t2s2 = null_or_int(s2, 1)
            t2s3 = null_or_int(s2, 2)

            lines.append(
                f"INSERT INTO matches"
                f" (season, week_label, played_on,"
                f"  team1_player1, team1_player2, team2_player1, team2_player2,"
                f"  team1_set1, team1_set2, team1_set3, team2_set1, team2_set2, team2_set3)"
                f" VALUES"
                f" ('{m['season']}', '{m['week_label']}', {date_str},"
                f"  {t1p1}, {t1p2}, {t2p1}, {t2p2},"
                f"  {t1s1}, {t1s2}, {t1s3}, {t2s1}, {t2s2}, {t2s3});"
            )
            total += 1
        lines.append('')
    lines.append(f'-- Total: {total} matches across {len(all_season_matches)} season(s)')
    return '\n'.join(lines)

# ── Filename helpers ──────────────────────────────────────────────────────────

def detect_season_from_filename(path):
    """Extract '2024-25' from 'WBP 24-25.xlsx' or 'WBP 10-11.xls'."""
    m = re.search(r'WBP (\d{2})-(\d{2})\.xlsx?', str(path), re.IGNORECASE)
    if m:
        y1, y2 = m.group(1), m.group(2)
        return f'20{y1}-{y2}'
    return None

def is_golf_file(path):
    """Return True if the file looks like a golf file."""
    name = os.path.basename(path).lower()
    return 'golf' in name

def sort_key(season_label):
    """Sort '2023-24' numerically."""
    m = re.match(r'(\d{4})-(\d{2})', season_label)
    if m:
        return int(m.group(1)) * 100 + int(m.group(2))
    return 0

# ── Main ─────────────────────────────────────────────────────────────────────

def process_file(path, season_label):
    """Process a single season file (.xlsx or .xls). Returns (matches, season_members, use_named_members)."""
    print(f'\nReading: {os.path.basename(path)}  →  season {season_label}')
    if path.lower().endswith('.xls'):
        return parse_xls_file(path, season_label)
    return parse_season_file(path, season_label)

def print_match_table(matches, season_members):
    """Print a formatted match list for review."""
    print(f'\n  {"#":>3}  {"Week":>6}  {"Court":8}  {"Team 1":22}  {"Score":10}  {"Team 2":22}  {"Score"}')
    print(f'  {"-"*90}')
    guests_seen = set()
    for i, m in enumerate(matches, 1):
        t1 = f"{m['team1_player1']}/{m['team1_player2']}"
        t2 = f"{m['team2_player1'] or '?'}/{m['team2_player2'] or '?'}"
        s1 = '/'.join(str(x) for x in m['team1_scores']) or '?'
        s2 = '/'.join(str(x) for x in m['team2_scores']) or '?'

        flags = []
        for name in [m['team1_player1'], m['team1_player2'], m['team2_player1'], m['team2_player2']]:
            if name and name not in season_members:
                guests_seen.add(name)
                flags.append(f'*{name}')
        flag_str = '  ' + ','.join(flags) if flags else ''
        print(f'  {i:3d}  {m["week_label"]:>6}  {m["court"]:8}  {t1:22}  {s1:10}  {t2:22}  {s2}{flag_str}')

    if guests_seen:
        print(f'\n  Guests flagged: {sorted(guests_seen)}')
    return guests_seen

def main():
    parser = argparse.ArgumentParser(description='Import WBP season Excel files into Supabase SQL.')
    group = parser.add_mutually_exclusive_group(required=True)
    group.add_argument('--file', help='Path to a single WBP season xlsx file')
    group.add_argument('--dir',  help='Directory containing WBP season xlsx files (processes all)')
    parser.add_argument('--season', help='Season label e.g. 2024-25 (auto-detected from filename if omitted; only for --file)')
    parser.add_argument('--skip',   help='Comma-separated season labels to skip e.g. 2024-25', default='')
    args = parser.parse_args()

    skip_seasons = {s.strip() for s in args.skip.split(',') if s.strip()}

    # ── Single file mode ──────────────────────────────────────────────────────
    if args.file:
        xlsx_path = args.file
        season_label = args.season or detect_season_from_filename(xlsx_path)
        if not season_label:
            print('ERROR: Could not detect season label from filename. Use --season 2024-25')
            sys.exit(1)

        matches, season_members, use_named = process_file(xlsx_path, season_label)
        print(f'\nMatches found: {len(matches)}')
        guests_seen = print_match_table(matches, season_members)

        # Collect new members to insert
        # In single-file mode: guests from named seasons, all from rated seasons
        new_members = []
        if use_named:
            for name in sorted(guests_seen):
                new_members.append((name, True))
        else:
            # All match participants need to be in members table
            for name in sorted(season_members):
                new_members.append((name, False))

        members_sql = generate_members_sql(new_members)
        data_sql    = generate_data_sql([(season_label, matches)])

        with open('historical-members.sql', 'w') as f:
            f.write(members_sql)
        print('\nWrote: historical-members.sql')

        with open('historical-data.sql', 'w') as f:
            f.write(data_sql)
        print('Wrote: historical-data.sql')

        print('\nNext steps:')
        print('  1. Review both SQL files for accuracy')
        print('  2. Run historical-members.sql in Supabase SQL editor')
        print('  3. Run historical-data.sql in Supabase SQL editor')

    # ── Directory (batch) mode ────────────────────────────────────────────────
    else:
        history_dir = args.dir
        # Collect both .xlsx and .xls files (glob *.xls does NOT match *.xlsx)
        all_files = sorted(
            glob.glob(os.path.join(history_dir, 'WBP *.xlsx')) +
            glob.glob(os.path.join(history_dir, 'WBP *.xls'))
        )

        season_files = []
        for path in all_files:
            if is_golf_file(path):
                print(f'  Skipping golf file: {os.path.basename(path)}')
                continue
            sl = detect_season_from_filename(path)
            if not sl:
                print(f'  Skipping (cannot parse season): {os.path.basename(path)}')
                continue
            if sl in skip_seasons:
                print(f'  Skipping (--skip): {sl}')
                continue
            season_files.append((path, sl))

        # Sort oldest → newest
        season_files.sort(key=lambda x: sort_key(x[1]))

        print(f'\nFound {len(season_files)} season file(s) to process:')
        for path, sl in season_files:
            print(f'  {sl}  →  {os.path.basename(path)}')

        # Track all unique players we'll need in the DB
        # key: name (lowercased for dedup), value: (canonical_name, is_guest)
        all_new_members = {}   # name.lower() -> (name, is_guest)
        all_season_data  = []  # list of (season_label, matches)

        for path, season_label in season_files:
            matches, season_members, use_named = process_file(path, season_label)
            print(f'  → {len(matches)} matches')

            all_season_data.append((season_label, matches))

            # Collect players to insert
            all_in_season = set()
            for m in matches:
                for name in [m['team1_player1'], m['team1_player2'],
                             m['team2_player1'], m['team2_player2']]:
                    if name:
                        all_in_season.add(name)

            if use_named:
                # Only guests need inserting (members assumed to already exist)
                guests = all_in_season - season_members
                for name in guests:
                    key = name.lower()
                    if key not in all_new_members:
                        all_new_members[key] = (name, True)   # is_guest
            else:
                # Pre-2017: all players need to be in members table
                for name in all_in_season:
                    key = name.lower()
                    if key not in all_new_members:
                        all_new_members[key] = (name, False)  # not a guest

        new_members_list = sorted(all_new_members.values(), key=lambda x: x[0])

        print(f'\n── Summary ──────────────────────────────────────────────')
        print(f'Seasons processed: {len(all_season_data)}')
        print(f'Total matches:     {sum(len(m) for _, m in all_season_data)}')
        print(f'New members/guests to insert: {len(new_members_list)}')
        if new_members_list:
            for name, is_guest in new_members_list:
                print(f'  {"[guest]" if is_guest else "[member]":9s} {name}')

        members_sql = generate_members_sql(new_members_list)
        data_sql    = generate_data_sql(all_season_data)

        with open('historical-members-all.sql', 'w') as f:
            f.write(members_sql)
        print('\nWrote: historical-members-all.sql')

        with open('historical-data-all.sql', 'w') as f:
            f.write(data_sql)
        print('Wrote: historical-data-all.sql')

        print('\nNext steps:')
        print('  1. Review both SQL files for accuracy')
        print('  2. Run historical-members-all.sql in Supabase SQL editor')
        print('  3. Run historical-data-all.sql in Supabase SQL editor')

if __name__ == '__main__':
    main()
